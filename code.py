from pathlib import Path
from openpyxl import load_workbook, Workbook

import tkinter as tk
from tkinter import filedialog, messagebox


def generar_listado_hojas(input_xlsx: str, output_xlsx: str) -> None:
    input_path = Path(input_xlsx)

    wb_in = load_workbook(input_path, read_only=False, data_only=True)

    visibles = []
    ocultas = []

    for i, ws in enumerate(wb_in.worksheets, start=1):
        estado = getattr(ws, "sheet_state", "visible")  # visible / hidden / veryHidden
        nombre = ws.title

        if estado == "visible":
            visibles.append((i, nombre, estado))
        else:
            ocultas.append((i, nombre, estado))

    wb_out = Workbook()
    ws_vis = wb_out.active
    ws_vis.title = "Hojas_visibles"
    ws_ocu = wb_out.create_sheet("Hojas_ocultas")

    headers = ["Orden", "Nombre_hoja", "Estado"]
    ws_vis.append(headers)
    ws_ocu.append(headers)

    for row in visibles:
        ws_vis.append(row)

    for row in ocultas:
        ws_ocu.append(row)

    for ws in (ws_vis, ws_ocu):
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15

    wb_out.save(output_xlsx)


def main():
    root = tk.Tk()
    root.withdraw()  # oculta la ventana principal
    root.update()

    input_file = filedialog.askopenfilename(
        title="Selecciona el archivo Excel de entrada",
        filetypes=[
            ("Excel (*.xlsx, *.xlsm)", "*.xlsx *.xlsm"),
            ("Todos los archivos", "*.*"),
        ],
    )

    if not input_file:
        messagebox.showinfo("Cancelado", "No seleccionaste ningún archivo.")
        return

    # Sugerir nombre de salida por defecto
    input_path = Path(input_file)
    default_name = f"listado_hojas_{input_path.stem}.xlsx"
    default_path = str(input_path.with_name(default_name))

    output_file = filedialog.asksaveasfilename(
        title="Guardar Excel de salida",
        defaultextension=".xlsx",
        initialfile=default_name,
        initialdir=str(input_path.parent),
        filetypes=[("Excel (*.xlsx)", "*.xlsx")],
    )

    if not output_file:
        messagebox.showinfo("Cancelado", "No seleccionaste dónde guardar el archivo de salida.")
        return

    try:
        generar_listado_hojas(input_file, output_file)
        messagebox.showinfo("Listo ✅", f"Archivo generado:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error:\n{e}")


if __name__ == "__main__":
    main()
